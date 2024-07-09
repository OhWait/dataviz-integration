import os
import re
import pandas as pd
import psycopg2
import shutil
from openpyxl import load_workbook
import xlrd
import tempfile
import numpy as np  # Ajouter l'importation de numpy pour gérer les valeurs non finies
from airflow import DAG
from airflow.operators.python_operator import PythonOperator
from airflow.utils.dates import days_ago

default_args = {
    'owner': 'airflow',
    'depends_on_past': False,
    'email_on_failure': False,
    'email_on_retry': False,
    'retries': 1
}

def get_folders():
    base_directory_path = '/opt/airflow/upload/territoire'
    folders = []

    for folder_name in os.listdir(base_directory_path):
        folder_path = os.path.join(base_directory_path, folder_name)

        if os.path.isdir(folder_path) and "intercommunalite" in folder_name.lower():
            folders.append(folder_path)
        elif os.path.isdir(folder_path) and "epci_" in folder_name.lower():
            folders.append(folder_path)
            
    return folders

def extract_year_from_folder(folder_name):
    pattern = r'.*(\d{4})'
    match = re.search(pattern, folder_name, re.IGNORECASE)
    if match:
        return match.group(1)
    return None

def find_header_row(sheet, required_columns):
    """
    Function to find the header row in a sheet that contains all required columns.
    """
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if all(col.strip().lower() in required_columns for col in row):
            return i
    return None

def load_excel(file_path, sheet_name, required_columns):
    """
    Function to load data from an Excel file (.xls or .xlsx) and map to required columns.
    """
    if file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_name(sheet_name)
        header_row = find_header_row_xls(sheet, required_columns)
        
        if header_row is None:
            raise ValueError(f"Required columns not found in {sheet_name} sheet of file {file_path}")
        
        data = [sheet.row_values(row) for row in range(header_row + 1, sheet.nrows)]
        df = pd.DataFrame(data, columns=[col.strip().lower() for col in sheet.row_values(header_row)])
    
    elif file_path.endswith('.xlsx'):
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        sheet = wb[sheet_name]
        header_row = find_header_row(sheet, required_columns)
        
        if header_row is None:
            raise ValueError(f"Required columns not found in {sheet_name} sheet of file {file_path}")
        
        data = sheet.iter_rows(min_row=header_row + 1, values_only=True)
        df = pd.DataFrame(data, columns=[col.strip().lower() for col in next(data)])
        
        wb.close()
    
    else:
        raise ValueError(f"Unsupported file format for {file_path}. Only .xls and .xlsx are supported.")
    
    return df

def find_header_row_xls(sheet, required_columns):
    """
    Function to find the header row in a sheet (for .xls files).
    """
    for i in range(sheet.nrows):
        if all(col.strip().lower() in required_columns for col in sheet.row_values(i)):
            return i
    return None

def process_epci_sheet(file_path, year):
    try:
        required_columns = ['epci', 'libepci', 'nature_epci', 'nb_com']
        df = load_excel(file_path, 'EPCI', required_columns)
        print(f"{file_path} : {df.head()}")
        
        df['annee'] = year
        df['nb_com'] = pd.to_numeric(df['nb_com'], errors='coerce')
        
        # Gérer les valeurs non finies
        nan_values = df['nb_com'].isna()
        if nan_values.any():
            print(f"Found NaN or non-numeric values in 'nb_com' column for file {file_path}:")
            print(df[nan_values])
        
        # Convertir les valeurs restantes en entier
        df['nb_com'] = df['nb_com'].fillna(0).astype(int)
        df = df.dropna(subset=required_columns)
        
        df = df[['annee', 'epci', 'libepci', 'nature_epci', 'nb_com']]
        
        insert_data(df, 'epci')
    except Exception as e:
        print(f"Failed to process EPCI sheet from file {file_path}: {e}")
        raise

def process_composition_communale_sheet(file_path, year):
    try:
        required_columns = ['codgeo', 'libgeo', 'epci', 'libepci', 'dep', 'reg']
        df = load_excel(file_path, 'Composition_communale', required_columns)
        print(f"{file_path} : {df.head()}")
        
        df['annee'] = year
        df = df.dropna(subset=required_columns)
        
        df = df[['annee', 'codgeo', 'libgeo', 'epci', 'libepci']]
        
        insert_data(df, 'epci_composition')
    except Exception as e:
        print(f"Failed to process Composition_communale sheet from file {file_path}: {e}")
        raise

def clean_table(table_name, year):
    try:
        conn = psycopg2.connect(
            dbname=os.getenv('POSTGRES_DB'),
            user=os.getenv('POSTGRES_USER'),
            password=os.getenv('POSTGRES_PASSWORD'),
            host=os.getenv('POSTGRES_HOST'),
            port=os.getenv('POSTGRES_PORT')
        )
        cursor = conn.cursor()
        
        delete_sql = f"DELETE FROM territoire.{table_name} WHERE annee = %s"
        cursor.execute(delete_sql, (year,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        print(f"Successfully cleaned data for year {year} in PostgreSQL table {table_name}.")
        
    except Exception as e:
        print(f"Failed to clean data for year {year} in PostgreSQL table {table_name}: {e}")
        raise

def insert_data(df, table_name):
    try:
        conn = psycopg2.connect(
            dbname=os.getenv('POSTGRES_DB'),
            user=os.getenv('POSTGRES_USER'),
            password=os.getenv('POSTGRES_PASSWORD'),
            host=os.getenv('POSTGRES_HOST'),
            port=os.getenv('POSTGRES_PORT')
        )
        cursor = conn.cursor()
    
        transformed_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
        df.to_csv(transformed_file.name, index=False, na_rep='')

        print(df.head())

        copy_sql = f"""
            COPY territoire.{table_name}
            FROM STDIN WITH CSV HEADER
            DELIMITER AS ','
            NULL AS ''
        """

        with open(transformed_file.name, 'r') as f:
            cursor.copy_expert(copy_sql, f)

        conn.commit()
        cursor.close()
        conn.close()
        print(f"Successfully inserted data into PostgreSQL table {table_name}.")
        
        # Suppression du fichier temporaire
        os.remove(transformed_file.name)
        print(f"Successfully deleted temporary file {transformed_file.name}.")
        
    except Exception as e:
        print(f"Failed to insert data into PostgreSQL table {table_name}: {e}")
        raise

def process_files_by_year(folder, year):
    files = [f for f in os.listdir(folder) if f.endswith(('.xls', '.xlsx'))]

    for file_name in files:
        file_path = os.path.join(folder, file_name)
        process_epci_sheet(file_path, year)
        process_composition_communale_sheet(file_path, year)

    # If no exception was raised, delete the folder
    try:
        shutil.rmtree(folder)
        print(f"Successfully deleted folder {folder}.")
    except Exception as e:
        print(f"Failed to delete folder {folder}: {e}")
        raise

with DAG(
    'epci_integration',
    default_args=default_args,
    description='Import EPCI and Composition_communale data from Excel files',
    schedule_interval=None,
    start_date=days_ago(1),
    catchup=False
) as dag:
    folders = get_folders()
    
    for folder in folders:
        year = extract_year_from_folder(folder)
        
        process_task = PythonOperator(
            task_id=f'process_files_{year}',
            python_callable=process_files_by_year,
            op_kwargs={'folder': folder, 'year': year}
        )
        process_task
